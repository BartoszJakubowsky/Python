import openpyxl
from openpyxl.styles import Alignment, Border, Side

import re
import math
import shutil
import xlwings
from pathlib import Path

mainCable = None;
calculatedPoles = []
def ReadExcelData(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    
    # Get the active sheet
    sheet = workbook.active
    
    # Get the number of rows and columns
    rows = sheet.max_row
    cols = sheet.max_column
    
    # Initialize the data table
    data_table = []

    # Iterate through rows, starting from row 1 and incrementing by 4
    for row_start in range(1, rows + 1, 4):
        # Initialize an object for the current range
        data_object = []
        
        # Iterate through columns
        for col in range(1, cols + 1):
            # Get values from the current range
            cell_value = [sheet.cell(row=row, column=col).value for row in range(row_start, row_start + 4)]
            
            # Add values to the object
            data_object.append(cell_value)
        
        # Add the object to the data table
        data_table.append(data_object)

    return data_table
#table of tables -> one table one pole with all equipment

def countLength(cordsA, cordsB):
    x1, y1 = cordsA
    x2, y2 = cordsB

    length = round(math.sqrt((x2 - x1)**2 + (y2 - y1)**2))

    return length
def countDeq(coordsA):
    global mainCable
    if (mainCable is None):
        return 90

    coordsB = mainCable

    vector1 = [b - a for a, b in zip(coordsA[0], coordsB[0])]
    vector2 = [b - a for a, b in zip(coordsA[1], coordsB[1])]

    dotProduct = sum(a * b for a, b in zip(vector1, vector2))
    vectorLength1 = math.sqrt(sum(a**2 for a in vector1))
    vectorLength2 = math.sqrt(sum(b**2 for b in vector2))

    if vectorLength1 == 0 or vectorLength2 == 0:
        deg = 180
    else:
        degCos = dotProduct / (vectorLength1 * vectorLength2)
        deg = round(math.degrees(math.acos(min(1, max(-1, degCos)))))

    return deg + 90
def formatCablesString(inputString):
    #trim string
    trimmed_string = inputString.strip()

    #separate satring after +
    parts = trimmed_string.split('+')
    parts = trimmed_string.split('-')

    #format string
    for i in range(len(parts)):
        part = parts[i]

        if 'adss' in part:
            last_s_index = part.rfind('s')
            parts[i] = part[:last_s_index + 1] + ' ' + part[last_s_index + 1:]

        if 'al' in part:
            if 'l.' not in part:
                parts[i] = part.replace('l', 'l. ')
            else:
                parts[i] = part.replace('l.', 'l. ')

        if 'asxsn' in part:
            parts[i] = part.replace('n', 'n ')

    return parts
def formatCoordsString(coordsA):

    coordsA = coordsA.strip('()')
    coordsA = coordsA.split(' ')
    x1, y1 = map(float, coordsA[:2])
 
    return [x1, y1];
def getPoleFromData(data):
    formattedString = data.strip('()')
    separatedStrings = formattedString.split(' ')

    pole, function, number = separatedStrings
    return {"pole":pole, "function": function, "number": number}
def handle_P(data, excel):
    indicator = data[0]
    #{"pole":pole, "function": function, "number": number}
    poleData = getPoleFromData(data[1])
    cords = data[3]
    
    print(poleData)
    excel['C78'] = poleData['number']; 
    excel['G78'] = poleData['pole'];
    excel['J78'] = poleData['function'].upper();
    excel['G88'] = 15; #mufa
    excel['L78'] = "-" #stacja
def handle_M(data, excel):
    global mainCable
    
    indicator = data[0].upper()
    cables = formatCablesString(data[1])
    coordsA = formatCoordsString(data[2])
    coordsB = formatCoordsString(data[3])

    rangeVectorA = [['E44', 'E45', 'E46', 'E47', 'E48', 'E49', 'E50', 'E51'],
                    ['H44', 'H45', 'H46', 'H47', 'H48', 'H49', 'H50', 'H51']]
    rangeVectorB = [['E52', 'E53', 'E54', 'E55', 'E56', 'E57', 'E58', 'E59'],
                    ['H52', 'H53', 'H54', 'H55', 'H56', 'H57', 'H58', 'H59']]
    
    rangeVectorSecondary = [['E65', 'E66', 'E67', 'E68', 'E69', 'E70', 'E71', 'E72', 'E73', 'E74', 'E75'],
                            ['H65', 'H66', 'H67', 'H68', 'H69', 'H70', 'H71', 'H72', 'H73', 'H74', 'H75']]

    def putValueToCell(cable, rangeVector):
        for i in range(len(rangeVector[0])):
            cellA = rangeVector[0][i]
            cellB = rangeVector[1][i]

            if (excel[cellA].value == None):
                cells = excel[cellA : cellB]                
                for cell in cells[0]:
                    cell.value = cable[0]
                    cable.pop(0)

                return

    for i in range(len(cables)):
        cableType = cables[i]
        cable = [cableType, None, countDeq([coordsA, coordsB]), countLength(coordsA, coordsB)]
        if (mainCable is None and indicator == "M"):
            putValueToCell(cable, rangeVectorA)
        else:
            if (indicator== "M"):
                putValueToCell(cable, rangeVectorB)
            else: 
                putValueToCell(cable, rangeVectorSecondary)

    #if this is first cable 
    if (mainCable is None and indicator == "M"):
        mainCable = [coordsA, coordsB]

def exportDataFromCalculatedExcel(excel):

    global calculatedPoles
    excel.active = excel["KALKULATOR"];
    sheet = excel.active;
    lp = len(calculatedPoles) + 1
    station = sheet['L78'].value.upper()
    number = sheet['C78'].value
    pole = sheet['G78'].value.upper()
    function = sheet['J78'].value.upper()
    
    sheetRange = [['C44', 'C45', 'C46', 'C47', 'C48', 'C49', 'C50', 'C51', 'C52', 'C53', 'C54', 'C55', 'C56', 'C57', 'C58', 'C59', 'C60', 'C61', 'C62', 'C63', 'C64', 'C65', 'C66', 'C67', 'C68', 'C69', 'C70', 'C71', 'C72', 'C73', 'C74', 'C75'],
                ['K44', 'K45', 'K46', 'K47', 'K48', 'K49', 'K50', 'K51', 'K52', 'K53', 'K54', 'K55', 'K56', 'K57', 'K58', 'K59', 'K60', 'K61', 'K62', 'K63', 'K64', 'K65', 'K66', 'K67', 'K68', 'K69', 'K70', 'K71', 'K72', 'K73', 'K74', 'K75']]
    
    dataFromSheetRange = []

    #for each row
    for i in range(len(sheetRange[0])):
        cellA = sheetRange[0][i]
        cellB = sheetRange[1][i]
        cells = sheet[cellA : cellB]                

        rowData = []
        for cell in cells[0]:
            rowData.append(cell.value)
        dataFromSheetRange.append(rowData)

    filterCell = 2 #3 column = cable type, if None == delete all array row
    filteredDataFromSheetRange = [row for row in dataFromSheetRange if row[filterCell] is not None]
    # filteredDataFromSheetRange = [[row for row in subRow if row is not None] for subRow in filteredDataFromSheetRange]
    filteredDataFromSheetRange = [[row for index, row in enumerate(eachRow) if index not in (1, 3)] for eachRow in filteredDataFromSheetRange]


    excel.active = excel[function]
    sheet = excel.active;
    
    maxX = round(float(sheet['D2'].value),2)
    maxY = round(float(sheet['D3'].value),2)
    realMaxX = round(maxX*0.1,2)
    realMaxY = round(maxY*0.1,2)
    calcX = round(float(sheet["B2"].value),2)
    calcY = round(float(sheet["B3"].value),2)
    addedX = round(float(sheet["G2"].value),2)
    addedY = round(float(sheet["G3"].value),2)

    pole = {
        "lp": lp,
        "station" : station,
        "number" : number,
        "pole" : pole,
        "function" : function, 
        "cables" : filteredDataFromSheetRange,
        "maxX" : maxX,
        "maxY" : maxY,
        "realMaxX": realMaxX,
        "realMaxY": realMaxY,
        "calcX" : calcX,
        "calcY" : calcY,
        "addedX" : addedX,
        "addedY" : addedY
    }

    calculatedPoles.append(pole)  

def cacheFormulasData(path):
    excel_app = xlwings.App(visible=False)
    excel_book = excel_app.books.open(path)
    excel_book.save()
    excel_book.close()
    excel_app.quit()

def handleExcelFile(excel, path):
    sourceExcel = "C:\\Users\\BFS\\Documents\\kalkulator.xlsx"
    
    if (excel is None):
        shutil.copy(sourceExcel, path)
        excel = openpyxl.load_workbook(path)
        return excel
    else: 
        excel.save(path)
        excel.close()
def handleData(data):

    for i in range(len(data)):
        global mainCable
        mainCable = None
        poleData = data[i]

        if (len(poleData) <=1): 
            print("error", poleData)
            return any
        

        newExcelPath = f'C:\\Users\\BFS\\Documents\\test{i}.xlsx'
        excel = handleExcelFile(None, newExcelPath)
        excel.active = excel["KALKULATOR"];
        sheet = excel.active;

        for j in range(len(poleData)):
            table = poleData[j]
            #for each table handle p, m, a
            indicator = table[0].upper()
            if (indicator == 'P'):
                handle_P(table, sheet)
                next
            if (indicator == 'M' or indicator == "A"):
                handle_M(table, sheet)
                next
        
        handleExcelFile(excel, newExcelPath)
        cacheFormulasData(newExcelPath)
        excel = openpyxl.load_workbook(newExcelPath, data_only=True)

        exportDataFromCalculatedExcel(excel)
        handleExcelFile(excel, newExcelPath)
    exportCalculatedData()

def exportCalculatedData():
    #check if file exist
    #open file
    global calculatedPoles
    # print(calculatedPoles[0]['cables'])
    # return
    path = "C:\\Users\\BFS\\Documents\\ZestawienieTemp.xlsx"
    def handleExcelForExportedData():
        file = Path(path)
        if (file.exists()):
            return openpyxl.load_workbook(path)
        else: 
            shutil.copy("C:\\Users\\BFS\\Documents\\Zestawienie obliczeń.xlsx", path)
            return openpyxl.load_workbook(path)
        
    excel = handleExcelForExportedData()
    sheet = excel.active;
    
    # excelRange = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
    excelCablesRange = ['F', 'G', 'H', 'I', 'J', 'K', 'L']
    excelRestCableDataRange = ['A','B','C','D','E','M','N','O','P','Q','R','S','T',"U"]
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row = 3;
    for calcPole in calculatedPoles:
        lp = calcPole["lp"]
        station = calcPole["station"]
        number = calcPole["number"]
        pole = calcPole["pole"]
        function = calcPole["function"] 
        cables = calcPole["cables"]
        maxX = calcPole["maxX"]
        maxY = calcPole["maxY"]
        realMaxX = calcPole["realMaxX"]
        realMaxY = calcPole["realMaxY"]
        calcX = calcPole["calcX"]
        calcY = calcPole["calcY"]
        addedX = calcPole["addedX"]
        addedY = calcPole["addedY"]

        sheet[f'A{row}'].value = lp
        sheet[f'B{row}'].value = station
        sheet[f'C{row}'].value = number
        sheet[f'D{row}'].value = pole
        sheet[f'E{row}'].value = function

        #catalog values
        sheet[f'M{row}'].value = maxX
        sheet[f'N{row}'].value = maxY

        #max * pole state
        sheet[f'O{row}'].value = realMaxX
        sheet[f'P{row}'].value = realMaxY

        #only electrical
        sheet[f'Q{row}'].value = calcX
        sheet[f'R{row}'].value = calcY

        sheet[f'S{row}'].value = function

        sheet[f'T{row}'].value = addedX
        sheet[f'U{row}'].value = addedY

        cableRow = row

        
        for cable in cables:
            for i in range(len(cable)):
                cableData = cable[i]
                column = excelCablesRange[i]
                cell = sheet[f'{column}{cableRow}']
                cell.value = cableData
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            cableRow += 1

        cableRow -= 1
        for colRange in excelRestCableDataRange:
            sheet.merge_cells(range_string = f'{colRange}{row}:{colRange}{cableRow}')
            # sheet.merge_cells(start_column = colRange ,start_row = row, end_column = colRange, end_row = cableRow)

        row = cableRow + 1

    handleExcelFile(excel, path)
result_table = ReadExcelData("C:\\Users\\BFS\\Documents\\polesData_wyniki.xlsx")
handleData(result_table)


